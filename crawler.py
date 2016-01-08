# -*- coding: utf-8 -*-
# coding=utf-8

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import string
import time
from BeautifulSoup import BeautifulSoup
from requests import Request, Session
from datetime import datetime
import opencc
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sTime = 0.1

excelFilename = 'vulsList.xlsx'
wb = load_workbook(excelFilename)

begin = datetime(2015, 12, 30)
end = datetime(2016, 1, 7)

AllCVEList = []
whiteList = []
blackList = []

print " = Read xlsx file = "

history = wb['run1105_1230']
nrows = history.max_row + 1

for i in range(2, nrows) :
	CVEs = str(history.cell(row = i, column = 5).value)
	if "," in CVEs : 
		AllCVEList = AllCVEList + CVEs.split(',')
	else :
		if (CVEs != "None") :
			AllCVEList.append(CVEs)

wl = wb['whiteList']
nrows = wl.max_row + 1

for i in range(2, nrows) :
	whiteList.append(str(wl.cell(row = i, column = 1).value))

bl = wb['blackList']
nrows = bl.max_row + 1

for i in range(2, nrows) :
	blackList.append(str(bl.cell(row = i, column = 1).value))

def debugInputInfo() :
	print 'blackList'
	print blackList
	print 'whiteList'
	print whiteList
	print 'AllCVEList'
	print AllCVEList

print " = Read xlsx file done = "

def getHttp(url) :
	time.sleep(sTime)
	s = Session()	
	req = Request('GET', url)
	prepped = req.prepare()
	prepped.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.99 Safari/537.36' 
	count = 0
	while 1:
		if (count >= 3) :
			return response
		print 'Get : %s' % url
		response = s.send(prepped)
		if (response.status_code == 200) :
			return response
		print response.status_code
		time.sleep(1)
		count += 1

def checkDate(begin, date, end) :
	if (begin <= date and date <= end) :
		return True
	return False

def inBlackList(title) :
	global blackList
	for black in blackList :
		if re.search(black, title, re.IGNORECASE) :
			return True
	return False

def checkCVE(CVE) :
	global AllCVEList
	if re.match('CVE-\d{4}-\d{4,7}', CVE) is None :
		return "NoCVE"
	if CVE in AllCVEList :
		return "CVErepeat"
	AllCVEList.append(CVE)
	return "New"

def filterCVEs(CVElist) :
	new = []
	for CVE in CVElist :
		status = checkCVE(CVE)
		if (status == "NoCVE") : 
			return status
		if (status == "New") :
			new.append(CVE)
	return new

def checkCVEs(CVElist) :
	if (CVElist == "NoCVE") :
		return "NoCVE"
	if (CVElist == []) :
		return "CVErepeat"
	return "New"

def inWhiteList(title) :
	global whiteList
	for white in whiteList :
		if re.search(white, title, re.IGNORECASE) :
			return white
	return False

def getRisk(cve) :
	r = getHttp("https://web.nvd.nist.gov/view/vuln/detail?vulnId=" + cve)
	contents = BeautifulSoup(r.content).find("div", {"id": "contents"})
	if re.search("cvss-detail", str(contents)) == None :
		return "Not Found"
	firstRow = contents.findAll("div", {"class" : "cvss-detail"})[0].find("div")
 	if "CVSS v3" in firstRow.getText() :
		firstRow = contents.findAll("div", {"class" : "cvss-detail"})[1].find("div")
	reRisk = re.search('\((.+?)\)', firstRow.getText())
	return reRisk.group(1)

def getRiskByCVElist(CVElist) :
	if (len(CVElist) == 1) :
		return getRisk(CVElist[0])
	riskMap = {"LOW" : 0, "MEDIUM" : 1, "HIGH" : 2}
	riskFlags = [0, 0, 0]
	for CVE in CVElist :
		risk = getRisk(CVE)
		if (risk == "Not Found") : 
			continue
		riskFlags[riskMap[risk]] = 1
		if (riskFlags[0] == 1 and riskFlags[2] == 1) :
			break
	riskMap = dict((value, key) for key, value in riskMap.iteritems())
	risksLen = riskFlags.count(1)
	if risksLen == 0 :
		return "Not Found"
	elif risksLen == 1 : 
		return riskMap[riskFlags.index(1)]
	elif risksLen == 2:
		output = ""
		first = 1
		for index, riskFlag in enumerate(riskFlags) :
			if (riskFlag == 1) :
				if (first == 1) :
					output = riskMap[index] + " ~ "
					first = 0
				else :
					output = output + riskMap[index]
		return output
	elif risksLen == 3 :
		return riskMap[0] + " ~ " + riskMap[2]
	return "ERROR"

def riskEn2Tw(risk) :
	risk = risk.replace("LOW", u"低")
	risk = risk.replace("MEDIUM", u"中")
	risk = risk.replace("HIGH", u"高")
	return risk

urlList = [
	'https://www.exploit-db.com/remote/?order_by=date&order=desc',
	'https://www.exploit-db.com/webapps/?order_by=date&order=desc',
	'https://www.exploit-db.com/local/?order_by=date&order=desc',
	'https://www.exploit-db.com/dos/?order_by=date&order=desc'
]

grayFill = PatternFill(start_color='FF969696', end_color='FF969696', fill_type='solid')
orangeFill = PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
def setData(data) :
	global line	
	for i in range(1, 8) :
		run.cell(row = line, column = i).value = data[i - 1]

	if (data[6] == "Black" or data[6] == "CVErepeat") :
		for i in range(1, 8) :
			run.cell(row = line, column = i).fill = grayFill
	elif (data[6] == "White" and data[4] == "") :
		for i in range(1, 8) :
			run.cell(row = line, column = i).fill = orangeFill
	elif (data[6] != "White" and line != 1) :
		for i in range(1, 8) :
			run.cell(row = line, column = i).fill = yellowFill
	line += 1

def getExploitDB(url) :
	r = getHttp(url)
	soup = BeautifulSoup(r.content)
	rows = soup.find("table").find("tbody").findAll("tr")
	date = 0
	for row in rows :
		cells = row.findAll("td")
		date = datetime.strptime(cells[0].getText(), "%Y-%m-%d")  
		title = cells[4].getText()
		platform = cells[5].getText()
		source = cells[4].find('a').get('href')	
		if (checkDate(begin, date, end) == False) :
			continue
		data = [date, title, platform, source, "", "", ""]
		if inBlackList(title) :
			data[6] = "Black"
			setData(data)
			continue
		sourceRequest = getHttp(source)
		sourceHttp = BeautifulSoup(sourceRequest.content)
		tdList = sourceHttp.find("table", {"class" : "exploit_list"}).findAll("td")
		aList = tdList[1].findAll("a")
	
		CVElist = "NoCVE"
		if (len(aList) != 0) :
			CVElistOrigin = re.findall('CVE-\d{4}-\d{4,7}', str(aList[0]))
			CVElist = filterCVEs(CVElistOrigin)
		
		data[6] = checkCVEs(CVElist)
		if (data[6] == "CVErepeat") : 
			data[4] = ",".join(CVElistOrigin)
			setData(data)
			continue
		if (data[6] == "New") :
			data[4] = ",".join(CVElist)
			data[5] = riskEn2Tw(getRiskByCVElist(CVElist))
		platform = inWhiteList(title)
		if (platform != False) :
			data[2] = platform
			data[6] = "White"
		setData(data)
	return date

hkcertURL = 'https://www.hkcert.org/security-bulletin?p_p_id=3tech_list_security_bulletin_full_WAR_3tech_list_security_bulletin_fullportlet&_3tech_list_security_bulletin_full_WAR_3tech_list_security_bulletin_fullportlet_cur='

def getHkcert(url) :
	r = getHttp(url)
	soup = BeautifulSoup(r.content)
	rows = soup.find("table", attrs={"class": "sdchk_table3"}).find("tbody").findAll("tr")
	date = 0
	for row in rows :
		cells = row.findAll("td")
		date = datetime.strptime(cells[3].getText(), "%Y / %m / %d")  
		title = cells[1].contents[0].getText()
		source = 'https://www.hkcert.org/' + str(cells[1].find('a').get('href'))
		if (checkDate(begin, date, end) == False) :
			continue
		data = [date, title, "", source, "", "", ""]
		if inBlackList(title) :
			data[6] = "Black"
			setData(data)
			continue
		sourceRequest = getHttp(source)
		sourceHttp = BeautifulSoup(sourceRequest.content)	
		content6 = sourceHttp.find("div", {"id" : "content6"})
		CVElist = "NoCVE"
		if (content6 != None) :
			liList = content6.findAll("li")
			CVElistOrigin = []
			for li in liList :
				CVElistOrigin.append(li.getText())
			CVElist = filterCVEs(CVElistOrigin)
		data[6] = checkCVEs(CVElist)
		if (data[6] == "CVErepeat") : 
			data[4] = ",".join(CVElistOrigin)
		if (data[6] != "New") : 
			setData(data)
			continue
		platform = inWhiteList(title)
		if (platform != False) :
			data[2] = platform
			data[6] = 'White'
		data[4] = ",".join(CVElist)
		data[5] = riskEn2Tw(getRiskByCVElist(CVElist))
		setData(data)
	return date

nsfocusURL = 'http://www.nsfocus.net/index.php?act=sec_bug'

def getNsfocus(url) :
	r = getHttp(url)
	r.encoding = r.apparent_encoding
	soup = BeautifulSoup(r.text)
	rows = soup.find("ul", attrs={"class": "vul_list"}).findAll("li")
	cc = opencc.OpenCC('s2t')
	date = 0
	for row in rows :
		# cn word print ERROR but save file OK
		date = datetime.strptime(row.find("span").getText(), "%Y-%m-%d")  
		# save utf8 tw use excel import OK
		title = cc.convert(row.find("a").getText())	
		source = "http://www.nsfocus.net" + str(row.find("a").get("href"))
		CVEnumber = ""
		CVEre = re.search('CVE-\d{4}-\d{4,7}', title)
		if (checkDate(begin, date, end) == False) :
			continue
		
		if (CVEre == None) :
			sourceRequest = getHttp(source)
			sourceHttp = BeautifulSoup(sourceRequest.content)
			CVEre = re.search('CVE-\d{4}-\d{4,7}', str(sourceHttp))
		if (CVEre != None) :	
			CVEnumber = CVEre.group(0)
		
		title = title.replace("(" + CVEnumber + ")", "")
		data = [date, title, "", source, CVEnumber, "", ""]
		if inBlackList(title) :
			data[6] = "Black"
			setData(data)
			continue
		data[6] = checkCVE(CVEnumber)
		if (data[6] == "CVErepeat") : 
			data[4] = CVEnumber
			setData(data)
			continue
		if (data[6] == "New") :
			data[4] = CVEnumber
			data[5] = riskEn2Tw(getRisk(CVEnumber))
		platform = inWhiteList(title)
		if (platform != False) :
			data[2] = platform
			data[6] = 'White'
		setData(data)
	return date

def crawlExploitDB() :
	for url in urlList :
		pg = 1;
		while(1) :
			pgdate = getExploitDB(url + "&pg=" + str(pg))	
			if (pgdate >= begin) :
				pg += 1
			else :
				break;

def crawlHkcert() :
	pg = 1;
	while(1) :
		pgdate = getHkcert(hkcertURL + str(pg))	
		if (pgdate >= begin) :
			pg += 1
		else :
			break;

def crawlNsfocus() :
	pg = 1;
	while(1) :
		pgdate = getNsfocus(nsfocusURL + "&page=" + str(pg))	
		if (pgdate >= begin) :
			pg += 1
		else :
			break;

line = 1
run = wb['run']
data = ['Date', 'Title', 'Platform', 'Source', 'CVE', 'Risk', 'Status']
setData(data)

crawlExploitDB()
crawlHkcert()
crawlNsfocus()

wb.save(excelFilename)
